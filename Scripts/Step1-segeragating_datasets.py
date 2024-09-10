import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from tqdm import tqdm
import argparse
import os

def process_step2(input_file, dash_threshold=0.2):
    df = pd.read_excel(input_file, sheet_name=0)

    a_count_index = df.columns.get_loc("A_Count")

    rows_to_drop = []
    for index, row in tqdm(df.iloc[0:].iterrows(), total=len(df), desc="Checking '-' percentage"):
        total_cells = a_count_index - 6
        dash_count = (row[5:a_count_index] == "-").sum()
        if dash_count / total_cells >= dash_threshold:
            rows_to_drop.append(index)

    df_cleaned = df.drop(rows_to_drop)

    rows_to_drop = []
    for index, row in tqdm(df_cleaned.iloc[0:].iterrows(), total=len(df_cleaned), desc="Checking deletion criteria"):
        f_a = row['f(A)']
        f_h = row['f(H)']
        if (f_a == 0 and f_h == 0) or (f_a == 1 and f_h == 0) or f_a > 0.9 or f_a < 0.1 or f_h < 0.15:
            rows_to_drop.append(index)

    df_final = df_cleaned.drop(rows_to_drop)

    wb = load_workbook(input_file)
    ws = wb.create_sheet("Segregating loci")

    for r_idx, row in tqdm(enumerate(dataframe_to_rows(df_final, index=False, header=True), 1), total=len(df_final)+1, desc="Writing to Excel"):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if value == 'A':
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif value == 'B':
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif value == 'H':
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    wb.save(input_file)

def process_step3(input_file, error_rate=0.03):
    df = pd.read_excel(input_file, sheet_name="Segregating loci")

    total_individuals = sum('Individual' in col for col in df.columns)
    error_threshold = total_individuals * error_rate
    df_filtered = df[((df['A_Count'] <= error_threshold) | (df['B_Count'] == 0)) | ((df['B_Count'] <= error_threshold) | (df['A_Count'] == 0))]

    for index, row in tqdm(df_filtered.iterrows(), total=df_filtered.shape[0], desc="Processing rows"):
        for col in df.columns[5:]:
            cell_value = row[col]
            if isinstance(cell_value, str):
                if row['A_Count'] <= error_threshold:
                    df_filtered.at[index, col] = cell_value.replace('A', 'B')
                if row['B_Count'] <= error_threshold:
                    df_filtered.at[index, col] = cell_value.replace('B', 'A')

    wb = load_workbook(input_file)
    ws = wb.create_sheet("1 to 1")

    for r_idx, row in tqdm(enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1), total=df_filtered.shape[0]+1, desc="Writing to Excel"):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if value == 'A':
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif value == 'B':
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif value == 'H':
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    wb.save(input_file)

def process_step4(input_file, threshold=0.055):
    df = pd.read_excel(input_file, sheet_name="Segregating loci")

    df_filtered = df[(df.iloc[:, 3] == 'H') & (df.iloc[:, 4] == 'H')]

    def check_ab_percentage(row):
        total_count = len(row) - row.isnull().sum() - 2
        a_count = (row == 'A').sum()
        b_count = (row == 'B').sum()
        if a_count < total_count * threshold or b_count < total_count * threshold:
            return False
        return True

    df_filtered = df_filtered[df_filtered.apply(check_ab_percentage, axis=1)]

    wb = load_workbook(input_file)
    ws = wb.create_sheet("1 to 2 to 1")

    for r_idx, row in tqdm(enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1), total=len(df_filtered)+1, desc="Writing to Excel"):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if value == 'A':
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif value == 'B':
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif value == 'H':
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    wb.save(input_file)

def process_parents(input_file):
    df = pd.read_excel(input_file, sheet_name="1 to 1")

    df_filtered1 = df[((df.iloc[:, 3] == 'H') & (df.iloc[:, 4] == 'A')) | ((df.iloc[:, 3] == 'H') & (df.iloc[:, 4] == 'B'))]
    df_filtered1.iloc[:, 5:] = df_filtered1.iloc[:, 5:].replace('B', 'A')

    wb = load_workbook(input_file)
    ws1 = wb.create_sheet("Parent1")

    for r_idx, row in tqdm(enumerate(dataframe_to_rows(df_filtered1, index=False, header=True), 1), total=len(df_filtered1)+1, desc="Writing to Parent1"):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            if value == 'A':
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif value == 'B':
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif value == 'H':
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    wb.save(input_file)

    df_filtered2 = df[((df.iloc[:, 4] == 'H') & (df.iloc[:, 3] == 'A')) | ((df.iloc[:, 4] == 'H') & (df.iloc[:, 3] == 'B'))]
    df_filtered2.iloc[:, 5:] = df_filtered2.iloc[:, 5:].replace('B', 'A')

    ws2 = wb.create_sheet("Parent2")

    for r_idx, row in tqdm(enumerate(dataframe_to_rows(df_filtered2, index=False, header=True), 1), total=len(df_filtered2)+1, desc="Writing to Parent2"):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            if value == 'A':
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif value == 'B':
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif value == 'H':
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    wb.save(input_file)

def main():
    parser = argparse.ArgumentParser(description='Process genetic data in Excel files.')
    
    parser.add_argument(
        '-i', '--input',
        required=True,
        help=(
            "Input Excel file. This file must be the result of step 0.5, "
            "where rows have been manually numbered and 'ref/alt' columns "
            "have been removed."
        )
    )
    
    parser.add_argument(
        '-d', '--dash-threshold',
        type=float,
        default=0.2,
        help=(
            "Threshold for the percentage of '-' allowed per marker. "
            "Markers with a missing value percentage higher than this threshold "
            "will be removed. "
            "Default value is 0.2, meaning 20 percent."
        )
    )
    
    parser.add_argument(
        '-e', '--error-rate',
        type=float,
        default=0.03,
        help=(
            "Threshold for correcting potentially erroneous genotypes (percentage). "
            "This is used in the `process_step3` function to allow correction "
            "of genotypes that might have been mistakenly identified. "
            "Setting this to 0 means no genotype correction will be accepted. "
            "Default value is 0.03, meaning 3 percent."
        )
    )
    
    parser.add_argument(
        '-t', '--threshold',
        type=float,
        default=0.055,
        help=(
            "The minimum percentage of homozygous genotypes (A or B) in the parental "
            "separation matrix required to retain a marker. Higher values impose stricter "
            "criteria. Used in `process_step4`. "
            "Default value is 0.055, meaning 5.5 percent."
        )
    )
    
    args = parser.parse_args()

    input_file = args.input
    dash_threshold = args.dash_threshold
    error_rate = args.error_rate
    threshold = args.threshold

    process_step2(input_file, dash_threshold)
    process_step3(input_file, error_rate)
    process_step4(input_file, threshold)
    process_parents(input_file)

if __name__ == "__main__":
    main()
