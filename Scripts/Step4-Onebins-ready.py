import pandas as pd
import openpyxl
import numpy as np
from openpyxl.styles import Border, Side
from tqdm import tqdm
import argparse
import os

# Define the border style for the top border
thin_border = Border(top=Side(style='thin'))

# Function to process individual columns and apply borders, ignoring '-'
def process_individual_column_corrected(df, column_name, ws):
    previous_value = None
    for i in tqdm(range(1, len(df)), desc=f"Processing {column_name}"):
        current_value = df.at[i, column_name]
        if previous_value is not None and previous_value != '-' and current_value != '-':
            if current_value != previous_value:
                for col in range(1, len(df.columns) + 1):
                    ws.cell(row=i + 2, column=col).border = thin_border
        if current_value != '-':
            previous_value = current_value

# Function to process a single bin and return the representative row
def process_bin(df, start_row, end_row, individual_columns):
    bin_df = df.iloc[start_row:end_row+1]
    representative_row = bin_df.iloc[0].copy()  # Start with the first row as representative
    
    for col in individual_columns:
        col_values = bin_df[col].values
        original_value = representative_row[col]  # Get the original value from the first row
        col_values = col_values[col_values != '-']  # Ignore '-'
        
        if len(col_values) > 0:
            most_common = np.bincount([ord(c.upper()) for c in col_values]).argmax()
            most_common_char = chr(most_common)
            # Use lowercase only if the most common value differs from the original value
            if original_value.upper() != most_common_char:
                representative_row[col] = most_common_char.lower()
            else:
                representative_row[col] = original_value
    
    return representative_row

# Main processing function
def process_file(input_file, output_file):
    # Load the workbook and check for 'Processed Data' sheet
    wb = openpyxl.load_workbook(input_file)
    sheet_names = wb.sheetnames
    if 'Processed Data' in sheet_names:
        ws = wb['Processed Data']
    else:
        ws = wb[sheet_names[0]]  # Use the first sheet if 'Processed Data' is not found

    # Load the data into a pandas dataframe
    df = pd.read_excel(input_file, sheet_name=ws.title)

    # Convert 'H' to 'h' and 'A' to 'a' in all 'Individual' columns, leaving '-' as is
    individual_columns = [col for col in df.columns if col.startswith('Individual')]
    for col in individual_columns:
        df[col] = df[col].replace({'H': 'h', 'A': 'a'})

    # Apply borders where necessary
    for col in individual_columns:
        process_individual_column_corrected(df, col, ws)

    # Save the modified workbook with borders added
    intermediate_output = f"{os.path.splitext(input_file)[0]}_intermediate.xlsx"
    wb.save(intermediate_output)

    # Reload the modified workbook to generate Onebins
    new_processed_data_df = pd.read_excel(intermediate_output, sheet_name='Processed Data')
    new_wb = openpyxl.load_workbook(intermediate_output)
    new_ws = new_wb['Processed Data']

    # Create a new sheet for the "Onebins"
    new_onebins_ws = new_wb.create_sheet('Onebins')

    # Copy the header from the original sheet to "Onebins"
    for col_num, header in enumerate(new_processed_data_df.columns, 1):
        new_onebins_ws.cell(row=1, column=col_num).value = header

    # Process each bin and store the results
    processed_bins = []
    start_row = 0
    for row in tqdm(range(len(new_processed_data_df)), desc="Processing bins"):
        if any(new_ws.cell(row=row + 2, column=col).border.top.style for col in range(1, len(new_processed_data_df.columns) + 1)):
            representative_row = process_bin(new_processed_data_df, start_row, row - 1, individual_columns)
            processed_bins.append(representative_row)
            start_row = row  # Start of the next bin

    # Process the last bin
    if start_row < len(new_processed_data_df):
        representative_row = process_bin(new_processed_data_df, start_row, len(new_processed_data_df) - 1, individual_columns)
        processed_bins.append(representative_row)

    # Write the processed bins to the "Onebins" sheet
    for row_num, row_data in enumerate(processed_bins, 2):
        for col_num, value in enumerate(row_data, 1):
            new_onebins_ws.cell(row=row_num, column=col_num).value = value

    # Save the final workbook
    new_wb.save(output_file)
    print(f"Processing complete. Output saved to {output_file}")

# Command-line interface
def main():
    parser = argparse.ArgumentParser(description="Process Excel data and create Onebins sheet.")
    parser.add_argument('-i', '--input', required=True, help="Path to the input Excel file.")
    parser.add_argument('-o', '--output', help="Path to the output Excel file. Default: <input_file>_onebins.xlsx")

    args = parser.parse_args()

    input_file = args.input
    output_file = args.output or f"{os.path.splitext(input_file)[0]}_onebins.xlsx"

    if not os.path.isfile(input_file):
        print(f"Error: File {input_file} does not exist.")
        return

    process_file(input_file, output_file)

if __name__ == "__main__":
    main()

